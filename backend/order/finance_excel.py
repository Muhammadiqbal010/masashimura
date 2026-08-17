"""
finance_excel.py
================
Ekspor laporan keuangan Masashimura ke Excel — dibangun sepenuhnya dari kode.

FITUR UTAMA:
  - Logo asli di-embed di Cover (PNG dari LOGO_PATH, fallback ke teks)
  - Tanggal generate OTOMATIS dari timezone.now() — tidak perlu di-hardcode
  - Mendukung mode BULANAN (rekap per hari) dan TAHUNAN (rekap per bulan)
  - Semua sheet: Cover, Analisis Tren, Rekap Metode Pembayaran,
                 Rekap Periode, Ringkasan, Detail Transaksi, Top Menu, Pengeluaran

Dipanggil dari view Django melalui export_finance_excel_view().
"""

import calendar
import io
import os

from django.db.models import Count, DecimalField, ExpressionWrapper, F, Sum
from django.db.models.functions import ExtractMonth, TruncDate
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from finance.models import Expense
from .models import Order, OrderItem

# ── Path ke logo ─────────────────────────────────────────────────────────────
try:
    from django.conf import settings
    LOGO_PATH = getattr(
        settings,
        "EXCEL_LOGO_PATH",
        os.path.join(settings.BASE_DIR, "assets", "masashimura-logo.png"),
    )
except Exception:
    LOGO_PATH = "assets/masashimura-logo.png"

# ── Nama bulan Bahasa Indonesia ───────────────────────────────────────────────
BULAN_ID = {
    1: "Januari",   2: "Februari",  3: "Maret",     4: "April",
    5: "Mei",       6: "Juni",      7: "Juli",       8: "Agustus",
    9: "September", 10: "Oktober",  11: "November",  12: "Desember",
}

# ── Number formats ────────────────────────────────────────────────────────────
FMT_RP    = "#,##0"
FMT_PCT   = "0.00%"
FMT_DATE  = "DD/MM/YYYY"
FMT_INT   = "#,##0"

# ── Color palette ─────────────────────────────────────────────────────────────
C_DARK      = "111827"
C_DARK2     = "1F2937"
C_DARK3     = "374151"
C_GRAY      = "6B7280"
C_LIGHT     = "F9FAFB"
C_STRIPE    = "F3F4F6"
C_WHITE     = "FFFFFF"
C_RED       = "CC0000"
C_GOLD      = "F59E0B"
C_GREEN     = "059669"
C_BLUE      = "2563EB"
C_ORANGE    = "D97706"
C_PURPLE    = "7C3AED"
C_TEAL      = "0D9488"
C_RED_LT    = "FEE2E2"
C_GRN_LT    = "D1FAE5"
C_YLW_LT    = "FEF3C7"
C_BLUE_LT   = "EFF6FF"
C_TEAL_LT   = "CCFBF1"

# ── Sheet names & tab colors ──────────────────────────────────────────────────
SH_COVER    = "🏠 Cover"
SH_TREN     = "📈 Analisis Tren"
SH_PAYMENT  = "💳 Metode Pembayaran"
SH_REKAP    = "📊 Rekap Periode"
SH_RING     = "📋 Ringkasan"
SH_DETAIL   = "📄 Detail Transaksi"
SH_TOPMENU  = "🍜 Top Menu"
SH_EXPENSE  = "💰 Pengeluaran"

TAB_COLORS = {
    SH_COVER:   C_DARK,
    SH_TREN:    C_BLUE,
    SH_PAYMENT: C_TEAL,
    SH_REKAP:   C_BLUE,
    SH_RING:    C_GREEN,
    SH_DETAIL:  C_ORANGE,
    SH_TOPMENU: "DC2626",
    SH_EXPENSE: C_PURPLE,
}


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER STYLES
# ═══════════════════════════════════════════════════════════════════════════════

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color=C_DARK, italic=False, name="Arial") -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_thin(color="E5E7EB") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _border_bottom(color="E5E7EB") -> Border:
    return Border(bottom=Side(style="thin", color=color))


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _set_row_height(ws, row: int, height: float):
    ws.row_dimensions[row].height = height


def _write(ws, row, col, val, bold=False, size=10, color=C_DARK,
           bg=None, italic=False, h="left", v="center", wrap=False,
           num_fmt=None, border=None):
    c = ws.cell(row, col, val)
    c.font = _font(bold, size, color, italic)
    if bg:
        c.fill = _fill(bg)
    c.alignment = _align(h, v, wrap)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = border
    return c


def _merge(ws, r1, c1, r2, c2, val="", bold=False, size=10,
           color=C_DARK, bg=None, italic=False, h="left", v="center"):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(r1, c1, val)
    c.font = _font(bold, size, color, italic)
    if bg:
        c.fill = _fill(bg)
    c.alignment = _align(h, v)
    return c


def _banner(ws, row, col_start, col_end, text, bg, height=50, size=16):
    _set_row_height(ws, row, height)
    _merge(ws, row, col_start, row, col_end, text,
           bold=True, size=size, color=C_WHITE, bg=bg, h="left", v="center")


def _section_header(ws, row, col_start, col_end, text, height=28):
    _set_row_height(ws, row, height)
    _merge(ws, row, col_start, row, col_end, "  " + text,
           bold=True, size=11, color=C_WHITE, bg=C_DARK3, h="left", v="center")


def _table_header(ws, row, headers, start_col=1, bg=C_DARK, fg=C_WHITE, height=24):
    _set_row_height(ws, row, height)
    for i, h in enumerate(headers):
        c = ws.cell(row, start_col + i, h)
        c.font = Font(bold=True, color=fg, name="Arial", size=10)
        c.fill = _fill(bg)
        c.alignment = _align("center")
        c.border = Border(bottom=Side(style="medium", color=bg))
    ws.freeze_panes = ws.cell(row + 1, 1).coordinate


def _map_status(s: str) -> str:
    return {"paid": "Lunas", "pending": "Pending", "unpaid": "Pending"}.get(s, s.capitalize())


def _map_source(s: str) -> str:
    return "Web" if s == "web" else "POS"


# ═══════════════════════════════════════════════════════════════════════════════
# TANGGAL GENERATE
# ═══════════════════════════════════════════════════════════════════════════════

def _now_label() -> str:
    HARI_ID = {
        "Monday": "Senin", "Tuesday": "Selasa", "Wednesday": "Rabu",
        "Thursday": "Kamis", "Friday": "Jumat",
        "Saturday": "Sabtu", "Sunday": "Minggu",
    }
    now = timezone.localtime(timezone.now())
    hari = HARI_ID.get(now.strftime("%A"), now.strftime("%A"))
    return (
        f"{hari}, {now.day} {BULAN_ID[now.month]} {now.year}"
        f" — {now.strftime('%H:%M')} WIB"
    )


def _short_date(dt) -> str:
    return f"{dt.day} {BULAN_ID[dt.month]} {dt.year}"


def _build_filename(mode: str, month: int | None, year: int) -> str:
    if mode == "monthly" and month:
        return f"Rekap Finance Masashimura {BULAN_ID[month]} {year}.xlsx"
    return f"Rekap Finance Masashimura Tahun {year}.xlsx"


# ═══════════════════════════════════════════════════════════════════════════════
# WORKBOOK FACTORY
# ═══════════════════════════════════════════════════════════════════════════════

def _create_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for name in [SH_COVER, SH_TREN, SH_PAYMENT, SH_REKAP, SH_RING,
                 SH_DETAIL, SH_TOPMENU, SH_EXPENSE]:
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = TAB_COLORS[name]
    return wb


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: 🏠 COVER
# ═══════════════════════════════════════════════════════════════════════════════

def _write_cover(wb, period_label: str, generated_at: str,
                 mode: str, year: int,
                 total_pendapatan: float, total_pengeluaran: float,
                 total_laba: float):
    ws = wb[SH_COVER]
    _set_col_widths(ws, {"A": 1.5, "B": 2, "C": 22, "D": 22, "E": 22, "F": 22, "G": 2, "H": 1.5})

    for r in range(1, 40):
        _set_row_height(ws, r, 18)
        for col in "ABCDEFGH":
            ws[f"{col}{r}"].fill = _fill(C_DARK)

    for r in (1, 2):
        _set_row_height(ws, r, 6)
        for col in "BCDEFG":
            ws[f"{col}{r}"].fill = _fill(C_RED)

    logo_written = False
    if os.path.isfile(LOGO_PATH):
        try:
            from PIL import Image as PILImage
            pil_img  = PILImage.open(LOGO_PATH)
            orig_w, orig_h = pil_img.size
            logo_w = 500
            logo_h = max(1, int(orig_h * logo_w / orig_w))
            pil_img = pil_img.resize((logo_w, logo_h), PILImage.LANCZOS)
            canvas_w = 616
            canvas_h = logo_h
            offset_x = (canvas_w - logo_w) // 2
            canvas = PILImage.new("RGB", (canvas_w, canvas_h), (17, 24, 39))
            if pil_img.mode == "P":
                pil_img = pil_img.convert("RGBA")
            if pil_img.mode in ("RGBA", "LA"):
                canvas.paste(pil_img, (offset_x, 0), mask=pil_img.split()[-1])
            else:
                canvas.paste(pil_img.convert("RGB"), (offset_x, 0))
            buf = io.BytesIO()
            canvas.save(buf, format="PNG")
            buf.seek(0)
            xl_img        = XLImage(buf)
            xl_img.width  = canvas_w
            xl_img.height = canvas_h
            xl_img.anchor = "C4"
            ws.add_image(xl_img)
            logo_rows = max(4, int(canvas_h / 20) + 1)
            row_h = max(20, int(canvas_h / logo_rows))
            for r in range(4, 4 + logo_rows):
                _set_row_height(ws, r, row_h)
                for col in "ABCDEFGH":
                    ws[f"{col}{r}"].fill = _fill(C_DARK)
            logo_written = True
        except Exception as e:
            print(f"[LOGO ERROR] {e}")

    if not logo_written:
        _set_row_height(ws, 4, 8)
        _set_row_height(ws, 5, 60)
        _set_row_height(ws, 6, 8)
        ws.merge_cells("C5:F5")
        fb = ws["C5"]
        fb.value = "MASASHIMURA"
        fb.font  = Font(bold=True, size=36, color=C_GOLD, name="Arial Black")
        fb.fill  = _fill(C_DARK)
        fb.alignment = _align("center", "center")

    AFTER_LOGO = 10
    _set_row_height(ws, AFTER_LOGO, 4)
    for col in "CDEF":
        ws[f"{col}{AFTER_LOGO}"].fill = _fill(C_GOLD)

    _set_row_height(ws, AFTER_LOGO + 1, 22)
    ws.merge_cells(f"C{AFTER_LOGO+1}:F{AFTER_LOGO+1}")
    tl = ws[f"C{AFTER_LOGO+1}"]
    tl.value = "Sistem Laporan Keuangan Otomatis"
    tl.font  = Font(italic=True, size=12, color=C_GRAY, name="Arial")
    tl.fill  = _fill(C_DARK)
    tl.alignment = _align("center")

    BOX = AFTER_LOGO + 3
    for r in range(BOX, BOX + 5):
        _set_row_height(ws, r, 8 if r in (BOX, BOX + 4) else 36)
        for col in "CDEF":
            ws[f"{col}{r}"].fill = _fill(C_DARK2)

    ws.merge_cells(f"C{BOX+1}:F{BOX+1}")
    rt = ws[f"C{BOX+1}"]
    rt.value = "LAPORAN KEUANGAN"
    rt.font  = Font(bold=True, size=22, color=C_WHITE, name="Arial")
    rt.fill  = _fill(C_DARK2)
    rt.alignment = _align("center")

    ws.merge_cells(f"C{BOX+2}:F{BOX+2}")
    rp = ws[f"C{BOX+2}"]
    rp.value = f"PERIODE: {period_label.upper()}"
    rp.font  = Font(bold=True, size=14, color=C_GOLD, name="Arial")
    rp.fill  = _fill(C_DARK2)
    rp.alignment = _align("center")

    CARD = BOX + 6
    _set_row_height(ws, CARD,     6)
    _set_row_height(ws, CARD + 1, 30)
    _set_row_height(ws, CARD + 2, 30)
    _set_row_height(ws, CARD + 3, 6)

    mode_label = "Bulanan" if mode == "monthly" else "Tahunan"
    cards = [
        ("C", "D", "MODE",  mode_label, C_BLUE),
        ("E", "F", "TAHUN", str(year),  C_GOLD),
    ]
    for c1, c2, lbl, val, accent in cards:
        for r in range(CARD, CARD + 4):
            ws[f"{c1}{r}"].fill = _fill(C_DARK2)
            ws[f"{c2}{r}"].fill = _fill(C_DARK2)
        ws[f"{c1}{CARD}"].fill = _fill(accent)
        ws[f"{c2}{CARD}"].fill = _fill(accent)
        ws.merge_cells(f"{c1}{CARD+1}:{c2}{CARD+1}")
        lc = ws[f"{c1}{CARD+1}"]
        lc.value = lbl
        lc.font  = Font(size=9, color=C_GRAY, name="Arial")
        lc.fill  = _fill(C_DARK2)
        lc.alignment = _align("center")
        ws.merge_cells(f"{c1}{CARD+2}:{c2}{CARD+2}")
        vc = ws[f"{c1}{CARD+2}"]
        vc.value = val
        vc.font  = Font(bold=True, size=14, color=C_WHITE, name="Arial")
        vc.fill  = _fill(C_DARK2)
        vc.alignment = _align("center")

    STAT = CARD + 5
    _section_header(ws, STAT, 3, 6, "Ringkasan Keuangan", height=24)

    stats = [
        ("Total Pendapatan (Lunas)", total_pendapatan, C_GREEN),
        ("Total Pengeluaran",        total_pengeluaran, C_RED),
        ("Laba / Rugi Bersih",       total_laba,       C_ORANGE if total_laba >= 0 else "DC2626"),
    ]
    for i, (lbl, val, clr) in enumerate(stats):
        r = STAT + 1 + i
        _set_row_height(ws, r, 24)
        ws.merge_cells(f"C{r}:D{r}")
        lc = ws[f"C{r}"]
        lc.value = lbl
        lc.font  = Font(size=10, color=C_GRAY, name="Arial")
        lc.fill  = _fill(C_DARK2)
        lc.alignment = _align("left")
        ws.merge_cells(f"E{r}:F{r}")
        vc = ws[f"E{r}"]
        vc.value = f"Rp {int(val):,}".replace(",", ".")
        vc.font  = Font(bold=True, size=11, color=clr, name="Arial")
        vc.fill  = _fill(C_DARK2)
        vc.alignment = _align("right")

    FOOT = STAT + len(stats) + 2
    _set_row_height(ws, FOOT, 4)
    for col in "BCDEFG":
        ws[f"{col}{FOOT}"].fill = _fill(C_GOLD)

    _set_row_height(ws, FOOT + 2, 20)
    ws.merge_cells(f"C{FOOT+2}:F{FOOT+2}")
    ft = ws[f"C{FOOT+2}"]
    ft.value = f"v1.0  ·  Digenerate: {generated_at}  ·  © Masashimura {year}"
    ft.font  = Font(italic=True, size=9, color=C_GRAY, name="Arial")
    ft.fill  = _fill(C_DARK)
    ft.alignment = _align("center")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: 📈 ANALISIS TREN
# ═══════════════════════════════════════════════════════════════════════════════

def _write_analisis_tren(wb, mode: str, month: int | None, year: int,
                          orders_qs, expenses_qs, generated_at: str):
    """
    Bandingkan performa periode ini vs periode sebelumnya.
    - Bulanan : bulan ini vs bulan lalu (rekap per hari → tren mingguan)
    - Tahunan : tahun ini vs tahun lalu (rekap per bulan)
    """
    ws = wb[SH_TREN]
    _set_col_widths(ws, {
        "A": 2, "B": 18, "C": 18, "D": 18, "E": 16, "F": 16, "G": 2
    })

    period_str = f"{BULAN_ID[month]} {year}" if mode == "monthly" else f"Tahun {year}"
    _banner(ws, 1, 2, 6, f"📈  ANALISIS TREN — {period_str.upper()}", bg=C_BLUE)

    _set_row_height(ws, 2, 22)
    ws.merge_cells("B2:F2")
    sub = ws["B2"]
    sub.value = f"Perbandingan periode ini vs periode sebelumnya  |  Digenerate: {generated_at}"
    sub.font  = _font(italic=True, size=10, color=C_GRAY)
    sub.fill  = _fill(C_LIGHT)
    sub.alignment = _align("left")

    # ── Tentukan periode sebelumnya ───────────────────────────────────────────
    if mode == "monthly":
        if month == 1:
            prev_month, prev_year = 12, year - 1
        else:
            prev_month, prev_year = month - 1, year

        prev_orders_qs   = Order.objects.filter(
            created_at__year=prev_year, created_at__month=prev_month
        )
        prev_expenses_qs = Expense.objects.filter(
            date__year=prev_year, date__month=prev_month
        )
        prev_label = f"{BULAN_ID[prev_month]} {prev_year}"
    else:
        prev_orders_qs   = Order.objects.filter(created_at__year=year - 1)
        prev_expenses_qs = Expense.objects.filter(date__year=year - 1)
        prev_label = f"Tahun {year - 1}"

    # ── Hitung total ringkasan ─────────────────────────────────────────────────
    def _totals(o_qs, e_qs):
        rev   = float(o_qs.filter(payment_status="paid").aggregate(t=Sum("total_price"))["t"] or 0)
        exp   = float(e_qs.aggregate(t=Sum("amount"))["t"] or 0)
        cnt   = o_qs.filter(payment_status="paid").count()
        avg   = rev / cnt if cnt else 0
        return rev, exp, rev - exp, cnt, avg

    cur_rev,  cur_exp,  cur_net,  cur_cnt,  cur_avg  = _totals(orders_qs,      expenses_qs)
    prev_rev, prev_exp, prev_net, prev_cnt, prev_avg = _totals(prev_orders_qs, prev_expenses_qs)

    def _delta(cur, prev):
        if prev == 0:
            return None
        return (cur - prev) / prev

    def _arrow(pct, invert=False):
        if pct is None:
            return "—"
        good = pct >= 0 if not invert else pct <= 0
        arrow = "▲" if pct >= 0 else "▼"
        return f"{arrow} {abs(pct)*100:.1f}%"

    def _delta_color(pct, invert=False):
        if pct is None:
            return C_GRAY
        good = pct >= 0 if not invert else pct <= 0
        return C_GREEN if good else "DC2626"

    # ── Kartu perbandingan KPI ─────────────────────────────────────────────────
    _set_row_height(ws, 3, 10)
    _section_header(ws, 4, 2, 6, "📊  Perbandingan KPI Utama")

    headers_kpi = ["Indikator", period_str, prev_label, "Δ Perubahan", "Tren"]
    _table_header(ws, 5, headers_kpi, start_col=2, height=26)
    ws.freeze_panes = None

    kpi_rows = [
        ("💰 Pendapatan (Lunas)", cur_rev,  prev_rev,  False, FMT_RP),
        ("🛒 Pengeluaran",        cur_exp,  prev_exp,  True,  FMT_RP),
        ("📈 Laba Bersih",        cur_net,  prev_net,  False, FMT_RP),
        ("🧾 Jumlah Order Lunas", cur_cnt,  prev_cnt,  False, FMT_INT),
        ("💵 Rata-rata per Order",cur_avg,  prev_avg,  False, FMT_RP),
    ]

    DATA_START = 6
    for i, (label, cur, prev, invert, fmt) in enumerate(kpi_rows):
        r   = DATA_START + i
        bg  = C_WHITE if i % 2 == 0 else C_STRIPE
        pct = _delta(cur, prev)
        _set_row_height(ws, r, 28)

        lc = ws.cell(r, 2, label)
        lc.font = _font(bold=True, size=10); lc.fill = _fill(bg)
        lc.alignment = _align("left"); lc.border = _border_thin()

        cc = ws.cell(r, 3, cur)
        cc.number_format = fmt
        cc.font = _font(bold=True, size=10, color=C_BLUE)
        cc.fill = _fill(bg); cc.alignment = _align("right"); cc.border = _border_thin()

        pc = ws.cell(r, 4, prev)
        pc.number_format = fmt
        pc.font = _font(size=10, color=C_GRAY, italic=True)
        pc.fill = _fill(bg); pc.alignment = _align("right"); pc.border = _border_thin()

        delta_val = (cur - prev) if prev != 0 else cur
        dc = ws.cell(r, 5, delta_val if fmt == FMT_RP else delta_val)
        dc.number_format = fmt
        clr_d = _delta_color(pct, invert)
        dc.font = _font(bold=True, size=10, color=clr_d)
        dc.fill = _fill(C_GRN_LT if clr_d == C_GREEN else (C_RED_LT if clr_d == "DC2626" else C_LIGHT))
        dc.alignment = _align("right"); dc.border = _border_thin()

        tc = ws.cell(r, 6, _arrow(pct, invert))
        tc.font = _font(bold=True, size=11, color=_delta_color(pct, invert))
        tc.fill = _fill(bg); tc.alignment = _align("center"); tc.border = _border_thin()

    # ── Tren per sub-periode (per hari / per bulan) ───────────────────────────
    TREN_HEADER = DATA_START + len(kpi_rows) + 2
    _set_row_height(ws, TREN_HEADER - 1, 10)
    col_label = "Tanggal" if mode == "monthly" else "Bulan"
    _section_header(ws, TREN_HEADER, 2, 6,
                    f"📅  Tren Pendapatan per {'Hari' if mode=='monthly' else 'Bulan'}"
                    f" — {period_str} vs {prev_label}")

    sub_headers = [col_label, f"Pendapatan {period_str}", f"Pendapatan {prev_label}", "Δ (Rp)", "Δ (%)"]
    _table_header(ws, TREN_HEADER + 1, sub_headers, start_col=2, height=24)

    if mode == "monthly":
        def _rev_map(o_qs, yr, mo):
            qs = (
                o_qs.filter(payment_status="paid")
                .annotate(day=TruncDate("created_at"))
                .values("day")
                .annotate(total=Sum("total_price"))
            )
            return {str(r["day"]): float(r["total"] or 0) for r in qs}

        cur_map  = _rev_map(orders_qs,      year,      month)
        prev_map = _rev_map(prev_orders_qs, prev_year, prev_month)

        days_cur  = calendar.monthrange(year,      month)[1]
        days_prev = calendar.monthrange(prev_year, prev_month)[1]
        max_days  = max(days_cur, days_prev)

        sub_rows = []
        for d in range(1, max_days + 1):
            cur_key  = f"{year}-{month:02d}-{d:02d}"      if d <= days_cur  else None
            prev_key = f"{prev_year}-{prev_month:02d}-{d:02d}" if d <= days_prev else None
            c_val = cur_map.get(cur_key, 0)  if cur_key  else 0
            p_val = prev_map.get(prev_key, 0) if prev_key else 0
            label = f"{d:02d}/{month:02d}"
            sub_rows.append((label, c_val, p_val))
    else:
        def _rev_map_yearly(o_qs):
            qs = (
                o_qs.filter(payment_status="paid")
                .annotate(m=ExtractMonth("created_at"))
                .values("m")
                .annotate(total=Sum("total_price"))
            )
            return {r["m"]: float(r["total"] or 0) for r in qs}

        cur_map  = _rev_map_yearly(orders_qs)
        prev_map = _rev_map_yearly(prev_orders_qs)
        sub_rows = [
            (BULAN_ID[m], cur_map.get(m, 0), prev_map.get(m, 0))
            for m in range(1, 13)
        ]

    TREN_DATA = TREN_HEADER + 2
    for i, (lbl, c_val, p_val) in enumerate(sub_rows):
        r  = TREN_DATA + i
        bg = C_STRIPE if i % 2 == 0 else C_WHITE
        _set_row_height(ws, r, 20)

        diff = c_val - p_val
        pct_diff = diff / p_val if p_val else None
        has_data = c_val > 0 or p_val > 0

        lc = ws.cell(r, 2, lbl)
        lc.font = _font(bold=has_data, color=C_BLUE if has_data else C_GRAY)
        lc.fill = _fill(C_BLUE_LT if has_data else bg)
        lc.alignment = _align("center"); lc.border = _border_bottom()

        cc = ws.cell(r, 3, c_val)
        cc.number_format = FMT_RP
        cc.font = _font(bold=has_data, color=C_DARK if c_val else "D1D5DB")
        cc.fill = _fill(C_BLUE_LT if has_data else bg)
        cc.alignment = _align("right"); cc.border = _border_bottom()

        pc = ws.cell(r, 4, p_val)
        pc.number_format = FMT_RP
        pc.font = _font(color=C_GRAY, italic=True, size=9)
        pc.fill = _fill(bg); pc.alignment = _align("right"); pc.border = _border_bottom()

        dc = ws.cell(r, 5, diff)
        dc.number_format = FMT_RP
        clr_d = C_GREEN if diff > 0 else ("DC2626" if diff < 0 else C_GRAY)
        dc.font = _font(bold=(diff != 0), color=clr_d)
        dc.fill = _fill(C_GRN_LT if diff > 0 else (C_RED_LT if diff < 0 else bg))
        dc.alignment = _align("right"); dc.border = _border_bottom()

        pct_str = f"{pct_diff*100:+.1f}%" if pct_diff is not None else "—"
        pc2 = ws.cell(r, 6, pct_str)
        pc2.font = _font(bold=(diff != 0), color=clr_d)
        pc2.fill = _fill(bg); pc2.alignment = _align("center"); pc2.border = _border_bottom()

    # TOTAL row tren
    total_tren_r = TREN_DATA + len(sub_rows)
    _set_row_height(ws, total_tren_r, 26)
    t_cur  = sum(r[1] for r in sub_rows)
    t_prev = sum(r[2] for r in sub_rows)
    t_diff = t_cur - t_prev
    t_pct  = t_diff / t_prev if t_prev else None

    for col_n in range(2, 7):
        ws.cell(total_tren_r, col_n).fill = _fill(C_DARK)
    ws.cell(total_tren_r, 2, "TOTAL").font = _font(bold=True, size=11, color=C_WHITE)
    ws.cell(total_tren_r, 2).alignment = _align("center")

    for col_n, val, fmt in [(3, t_cur, FMT_RP), (4, t_prev, FMT_RP), (5, t_diff, FMT_RP)]:
        c = ws.cell(total_tren_r, col_n, val)
        c.number_format = fmt
        c.font = _font(bold=True, size=11, color=C_GOLD)
        c.fill = _fill(C_DARK); c.alignment = _align("right")

    pct_tot = f"{t_pct*100:+.1f}%" if t_pct is not None else "—"
    ws.cell(total_tren_r, 6, pct_tot).font = _font(bold=True, size=11,
        color=C_GREEN if t_diff >= 0 else "DC2626")
    ws.cell(total_tren_r, 6).fill = _fill(C_DARK)
    ws.cell(total_tren_r, 6).alignment = _align("center")

    # ── Insight otomatis ──────────────────────────────────────────────────────
    INS_R = total_tren_r + 2
    _section_header(ws, INS_R, 2, 6, "💡  Insight Otomatis")

    pct_rev = _delta(cur_rev, prev_rev)
    pct_exp = _delta(cur_exp, prev_exp)
    pct_net = _delta(cur_net, prev_net)

    def _rp(v): return f"Rp {int(v):,}".replace(",", ".")

    insights = []
    if pct_rev is not None:
        arah = "naik" if pct_rev >= 0 else "turun"
        insights.append(
            f"📊 Pendapatan {arah} {abs(pct_rev)*100:.1f}% vs {prev_label} "
            f"({_rp(prev_rev)} → {_rp(cur_rev)})"
        )
    if pct_exp is not None:
        arah = "naik" if pct_exp >= 0 else "turun"
        sikon = "⚠️" if pct_exp > 0.1 else "✅"
        insights.append(
            f"{sikon} Pengeluaran {arah} {abs(pct_exp)*100:.1f}% vs {prev_label} "
            f"({_rp(prev_exp)} → {_rp(cur_exp)})"
        )
    if pct_net is not None:
        arah = "meningkat" if pct_net >= 0 else "menurun"
        emoji = "📈" if pct_net >= 0 else "📉"
        insights.append(
            f"{emoji} Laba bersih {arah} {abs(pct_net)*100:.1f}% vs {prev_label} "
            f"({_rp(prev_net)} → {_rp(cur_net)})"
        )

    if sub_rows:
        best_day  = max(sub_rows, key=lambda x: x[1])
        worst_day = min((r for r in sub_rows if r[1] > 0), key=lambda x: x[1], default=None)
        insights.append(f"🏆 Hari/periode terbaik: {best_day[0]} — {_rp(best_day[1])}")
        if worst_day:
            insights.append(f"📌 Hari/periode terendah (ada transaksi): {worst_day[0]} — {_rp(worst_day[1])}")

    for j, text in enumerate(insights):
        r  = INS_R + 1 + j
        _set_row_height(ws, r, 22)
        ws.merge_cells(f"B{r}:F{r}")
        c = ws[f"B{r}"]
        c.value = "  " + text
        c.font  = _font(size=10)
        c.fill  = _fill(C_LIGHT if j % 2 == 0 else C_WHITE)
        c.alignment = _align("left")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: 💳 REKAP METODE PEMBAYARAN
# ═══════════════════════════════════════════════════════════════════════════════

def _write_rekap_metode_pembayaran(wb, orders_qs, period_label: str, generated_at: str):
    """
    Breakdown transaksi berdasarkan metode pembayaran:
    Cash, QRIS, Transfer/Gateway, dll.
    Tampilkan dominasi, tren per waktu, dan insight.
    """
    ws = wb[SH_PAYMENT]
    _set_col_widths(ws, {
        "A": 2, "B": 22, "C": 18, "D": 16, "E": 14, "F": 14, "G": 2
    })

    _banner(ws, 1, 2, 6, "💳  REKAP METODE PEMBAYARAN", bg=C_TEAL)

    _set_row_height(ws, 2, 22)
    ws.merge_cells("B2:F2")
    sub = ws["B2"]
    sub.value = f"Periode: {period_label}  |  Digenerate: {generated_at}"
    sub.font  = _font(italic=True, size=10, color=C_GRAY)
    sub.fill  = _fill(C_LIGHT)
    sub.alignment = _align("left")

    # ── Agregasi per metode ───────────────────────────────────────────────────
    raw = (
        orders_qs
        .values("payment_method", "payment_status")
        .annotate(
            jumlah=Count("id"),
            omzet=Sum("total_price"),
        )
        .order_by("payment_method")
    )

    # Kelompokkan: normalisasi nama metode
    METHOD_LABEL = {
        "cash":     "Cash",
        "qris":     "QRIS",
        "transfer": "Transfer",
        "gateway":  "Gateway / Online",
        "online":   "Gateway / Online",
        "debit":    "Kartu Debit/Kredit",
        "credit":   "Kartu Debit/Kredit",
    }
    METHOD_COLOR = {
        "Cash":               (C_GRN_LT,  C_GREEN),
        "QRIS":               (C_BLUE_LT, C_BLUE),
        "Transfer":           (C_YLW_LT,  C_ORANGE),
        "Gateway / Online":   (C_TEAL_LT, C_TEAL),
        "Kartu Debit/Kredit": ("F3E8FF",  C_PURPLE),
    }

    method_agg: dict[str, dict] = {}
    for row in raw:
        raw_m  = (row["payment_method"] or "cash").lower().strip()
        label  = METHOD_LABEL.get(raw_m, raw_m.capitalize())
        status = row["payment_status"]
        cnt    = row["jumlah"] or 0
        omz    = float(row["omzet"] or 0)

        if label not in method_agg:
            method_agg[label] = {"total_cnt": 0, "paid_cnt": 0, "total_omzet": 0, "paid_omzet": 0}
        method_agg[label]["total_cnt"]   += cnt
        method_agg[label]["total_omzet"] += omz
        if status == "paid":
            method_agg[label]["paid_cnt"]   += cnt
            method_agg[label]["paid_omzet"] += omz

    total_omzet_paid = sum(v["paid_omzet"] for v in method_agg.values())
    total_cnt_all    = sum(v["total_cnt"]   for v in method_agg.values())

    # Urutkan: dominan duluan
    sorted_methods = sorted(method_agg.items(), key=lambda x: x[1]["paid_omzet"], reverse=True)

    # ── Tabel utama ───────────────────────────────────────────────────────────
    _set_row_height(ws, 3, 10)
    _section_header(ws, 4, 2, 6, "📊  Ringkasan per Metode Pembayaran")

    headers = ["Metode", "Omzet Lunas (Rp)", "Jumlah Transaksi", "% Kontribusi", "Dominasi"]
    _table_header(ws, 5, headers, start_col=2, height=26)

    DATA_START = 6
    for i, (method, agg) in enumerate(sorted_methods):
        r  = DATA_START + i
        bg, fg = METHOD_COLOR.get(method, (C_STRIPE, C_DARK))
        _set_row_height(ws, r, 28)

        pct_omzet = agg["paid_omzet"] / total_omzet_paid if total_omzet_paid else 0
        is_top    = i == 0

        # Bar dominasi: 5 blok proporsional
        bar_filled = round(pct_omzet * 5)
        bar_str    = "█" * bar_filled + "░" * (5 - bar_filled)

        lc = ws.cell(r, 2, method)
        lc.font = _font(bold=is_top, size=11 if is_top else 10, color=fg)
        lc.fill = _fill(bg); lc.alignment = _align("left"); lc.border = _border_thin()

        oc = ws.cell(r, 3, agg["paid_omzet"])
        oc.number_format = FMT_RP
        oc.font = _font(bold=is_top, size=11 if is_top else 10, color=C_DARK)
        oc.fill = _fill(bg); oc.alignment = _align("right"); oc.border = _border_thin()

        cc = ws.cell(r, 4, agg["total_cnt"])
        cc.number_format = FMT_INT
        cc.font = _font(size=10, color=C_DARK)
        cc.fill = _fill(bg); cc.alignment = _align("center"); cc.border = _border_thin()

        pcc = ws.cell(r, 5, pct_omzet)
        pcc.number_format = FMT_PCT
        pcc.font = _font(bold=is_top, size=10, color=fg)
        pcc.fill = _fill(bg); pcc.alignment = _align("right"); pcc.border = _border_thin()

        bc = ws.cell(r, 6, bar_str + (" ◀ DOMINAN" if is_top else ""))
        bc.font = _font(bold=is_top, size=9,
                        color=fg if is_top else C_GRAY)
        bc.fill = _fill(bg); bc.alignment = _align("left"); bc.border = _border_thin()

    # TOTAL
    total_r = DATA_START + len(sorted_methods)
    _set_row_height(ws, total_r, 28)
    for col in range(2, 7):
        ws.cell(total_r, col).fill = _fill(C_DARK)
    ws.cell(total_r, 2, "TOTAL").font = _font(bold=True, size=11, color=C_WHITE)
    ws.cell(total_r, 2).alignment = _align("center")

    tc = ws.cell(total_r, 3, total_omzet_paid)
    tc.number_format = FMT_RP
    tc.font = _font(bold=True, size=11, color=C_GOLD)
    tc.fill = _fill(C_DARK); tc.alignment = _align("right")

    cntc = ws.cell(total_r, 4, total_cnt_all)
    cntc.number_format = FMT_INT
    cntc.font = _font(bold=True, size=11, color=C_GOLD)
    cntc.fill = _fill(C_DARK); cntc.alignment = _align("center")

    ws.cell(total_r, 5, 1.0).number_format = FMT_PCT
    ws.cell(total_r, 5).font = _font(bold=True, size=11, color=C_GOLD)
    ws.cell(total_r, 5).fill = _fill(C_DARK); ws.cell(total_r, 5).alignment = _align("right")

    # ── KPI Cards: metode dominan & detail ───────────────────────────────────
    CARD_R = total_r + 3
    _section_header(ws, CARD_R, 2, 6, "🏆  Metode Paling Dominan")
    _set_row_height(ws, CARD_R, 24)

    if sorted_methods:
        top_m, top_agg = sorted_methods[0]
        bg_top, fg_top = METHOD_COLOR.get(top_m, (C_STRIPE, C_DARK))
        pct_top = top_agg["paid_omzet"] / total_omzet_paid if total_omzet_paid else 0

        rows_card = [
            ("Metode Terdominan",   top_m,                                         fg_top),
            ("Omzet Lunas",         f"Rp {int(top_agg['paid_omzet']):,}".replace(",", "."),  C_GREEN),
            ("Jumlah Transaksi",    f"{top_agg['total_cnt']} transaksi",            C_BLUE),
            ("Kontribusi Omzet",    f"{pct_top*100:.1f}%",                         C_GOLD),
        ]
        for j, (lbl, val, clr) in enumerate(rows_card):
            r = CARD_R + 1 + j
            _set_row_height(ws, r, 26)
            bg = bg_top if j == 0 else (C_WHITE if j % 2 == 0 else C_LIGHT)

            ws.merge_cells(f"B{r}:C{r}")
            lc = ws[f"B{r}"]
            lc.value = lbl; lc.font = _font(size=10, color=C_GRAY)
            lc.fill  = _fill(bg); lc.alignment = _align("left"); lc.border = _border_thin()

            ws.merge_cells(f"D{r}:F{r}")
            vc = ws[f"D{r}"]
            vc.value = val
            vc.font  = _font(bold=True, size=11 if j == 0 else 10, color=clr)
            vc.fill  = _fill(bg); vc.alignment = _align("right"); vc.border = _border_thin()

    # ── Split: Cash vs Non-Cash ───────────────────────────────────────────────
    SPLIT_R = CARD_R + 7
    _section_header(ws, SPLIT_R, 2, 6, "💵  Split Cash vs Non-Cash")

    cash_omzet = sum(v["paid_omzet"] for k, v in method_agg.items() if k == "Cash")
    noncash_omzet = total_omzet_paid - cash_omzet
    cash_cnt   = sum(v["total_cnt"] for k, v in method_agg.items() if k == "Cash")
    noncash_cnt = total_cnt_all - cash_cnt

    pct_cash    = cash_omzet    / total_omzet_paid if total_omzet_paid else 0
    pct_noncash = noncash_omzet / total_omzet_paid if total_omzet_paid else 0

    split_rows = [
        ("💵 Cash",     cash_omzet,    cash_cnt,    pct_cash,    C_GRN_LT,  C_GREEN),
        ("📱 Non-Cash", noncash_omzet, noncash_cnt, pct_noncash, C_BLUE_LT, C_BLUE),
    ]
    split_headers = ["Tipe", "Omzet Lunas (Rp)", "Jumlah Transaksi", "% Omzet"]
    _table_header(ws, SPLIT_R + 1, split_headers, start_col=2, height=22,
                  bg=C_DARK3)

    for i, (lbl, omz, cnt, pct, bg, fg) in enumerate(split_rows):
        r = SPLIT_R + 2 + i
        _set_row_height(ws, r, 26)

        lc = ws.cell(r, 2, lbl)
        lc.font = _font(bold=True, size=11, color=fg)
        lc.fill = _fill(bg); lc.alignment = _align("left"); lc.border = _border_thin()

        oc = ws.cell(r, 3, omz)
        oc.number_format = FMT_RP
        oc.font = _font(bold=True, size=10, color=C_DARK)
        oc.fill = _fill(bg); oc.alignment = _align("right"); oc.border = _border_thin()

        cc = ws.cell(r, 4, cnt)
        cc.number_format = FMT_INT
        cc.font = _font(size=10); cc.fill = _fill(bg)
        cc.alignment = _align("center"); cc.border = _border_thin()

        pc = ws.cell(r, 5, pct)
        pc.number_format = FMT_PCT
        pc.font = _font(bold=True, size=10, color=fg)
        pc.fill = _fill(bg); pc.alignment = _align("right"); pc.border = _border_thin()

        # Mini bar visual
        filled = round(pct * 10)
        bar    = "█" * filled + "░" * (10 - filled)
        bc = ws.cell(r, 6, bar)
        bc.font = _font(size=9, color=fg); bc.fill = _fill(bg)
        bc.alignment = _align("left"); bc.border = _border_thin()

    # ── Insight ───────────────────────────────────────────────────────────────
    INS_R = SPLIT_R + 5
    _section_header(ws, INS_R, 2, 6, "💡  Insight Pembayaran")

    insights_pay = []
    if sorted_methods:
        top_m, top_agg = sorted_methods[0]
        pct_top = top_agg["paid_omzet"] / total_omzet_paid if total_omzet_paid else 0
        insights_pay.append(
            f"🏆 Metode paling dominan: {top_m} "
            f"({pct_top*100:.1f}% dari total omzet lunas)"
        )
    if total_omzet_paid > 0:
        if pct_cash >= 0.5:
            insights_pay.append(
                f"💵 Mayoritas transaksi masih tunai ({pct_cash*100:.1f}%) — "
                "pertimbangkan mendorong QRIS untuk efisiensi."
            )
        elif pct_noncash >= 0.6:
            insights_pay.append(
                f"📱 Non-cash sudah dominan ({pct_noncash*100:.1f}%) — "
                "transaksi digital makin tinggi."
            )
    if len(sorted_methods) > 1:
        bot_m, bot_agg = sorted_methods[-1]
        insights_pay.append(
            f"📉 Metode paling jarang digunakan: {bot_m} "
            f"({bot_agg['total_cnt']} transaksi)"
        )
    insights_pay.append(
        f"📊 Total {total_cnt_all} transaksi  |  "
        f"Omzet Lunas: Rp {int(total_omzet_paid):,}".replace(",", ".")
    )

    for j, text in enumerate(insights_pay):
        r = INS_R + 1 + j
        _set_row_height(ws, r, 22)
        ws.merge_cells(f"B{r}:F{r}")
        c = ws[f"B{r}"]
        c.value = "  " + text
        c.font  = _font(size=10)
        c.fill  = _fill(C_TEAL_LT if j % 2 == 0 else C_WHITE)
        c.alignment = _align("left")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: 📊 REKAP PERIODE
# ═══════════════════════════════════════════════════════════════════════════════

def _write_rekap_periode(wb, mode: str, month: int | None, year: int,
                          orders_qs, expenses_qs, generated_at: str):
    ws = wb[SH_REKAP]
    _set_col_widths(ws, {"A": 2, "B": 18, "C": 22, "D": 22, "E": 22, "F": 2})

    period_str = f"{BULAN_ID[month]} {year}" if mode == "monthly" else f"Tahun {year}"
    _banner(ws, 1, 2, 5, f"📊  REKAP PERIODE — {period_str.upper()}", bg=C_BLUE)

    _set_row_height(ws, 2, 22)
    ws.merge_cells("B2:E2")
    sub = ws["B2"]
    sub.value = (
        f"Mode: {'Bulanan (per hari)' if mode == 'monthly' else 'Tahunan (per bulan)'}"
        f"  |  Digenerate: {generated_at}"
    )
    sub.font  = _font(italic=True, size=10, color=C_GRAY)
    sub.fill  = _fill(C_LIGHT)
    sub.alignment = _align("left")

    _set_row_height(ws, 3, 8)

    col_label = "Tanggal" if mode == "monthly" else "Bulan"
    headers = [col_label, "Pendapatan (Rp)", "Pengeluaran (Rp)", "Laba Bersih (Rp)"]
    _table_header(ws, 4, headers, start_col=2, height=28)
    ws.freeze_panes = "B5"

    DATA_START = 5

    if mode == "monthly":
        rev_qs = (
            orders_qs.filter(payment_status="paid")
            .annotate(day=TruncDate("created_at"))
            .values("day")
            .annotate(total=Sum("total_price"))
            .order_by("day")
        )
        exp_qs = (
            expenses_qs
            .values("date")
            .annotate(total=Sum("amount"))
        )
        rev_map = {str(r["day"]): float(r["total"] or 0) for r in rev_qs}
        exp_map = {str(e["date"]): float(e["total"] or 0) for e in exp_qs}

        days_in_month = calendar.monthrange(year, month)[1]
        rows = []
        for d in range(1, days_in_month + 1):
            date_str  = f"{year}-{month:02d}-{d:02d}"
            date_disp = f"{d:02d}/{month:02d}/{year}"
            rev = rev_map.get(date_str, 0)
            exp = exp_map.get(date_str, 0)
            rows.append((date_disp, rev, exp, rev - exp))
    else:
        rev_qs = (
            orders_qs.filter(payment_status="paid")
            .annotate(m=ExtractMonth("created_at"))
            .values("m")
            .annotate(total=Sum("total_price"))
            .order_by("m")
        )
        exp_qs = (
            expenses_qs
            .annotate(m=ExtractMonth("date"))
            .values("m")
            .annotate(total=Sum("amount"))
        )
        rev_map = {r["m"]: float(r["total"] or 0) for r in rev_qs}
        exp_map = {e["m"]: float(e["total"] or 0) for e in exp_qs}
        rows = [
            (BULAN_ID[m], rev_map.get(m, 0), exp_map.get(m, 0),
             rev_map.get(m, 0) - exp_map.get(m, 0))
            for m in range(1, 13)
        ]

    for i, (label, rev, exp, net) in enumerate(rows):
        r  = DATA_START + i
        bg = C_STRIPE if i % 2 == 0 else C_WHITE
        has_data = rev > 0 or exp > 0
        _set_row_height(ws, r, 20)

        lc = ws.cell(r, 2, label)
        lc.font      = _font(bold=has_data, color=C_BLUE if has_data else C_DARK)
        lc.fill      = _fill(C_BLUE_LT if has_data else bg)
        lc.alignment = _align("center")

        for ci, val in [(3, rev), (4, exp), (5, net)]:
            c = ws.cell(r, ci, val)
            c.number_format = FMT_RP
            c.alignment     = _align("right")
            if val == 0:
                c.font = _font(color="D1D5DB")
                c.fill = _fill(C_BLUE_LT if has_data else bg)
            elif ci == 5 and val < 0:
                c.font = _font(bold=True, color="DC2626")
                c.fill = _fill(C_RED_LT)
            elif ci == 5 and val > 0:
                c.font = _font(bold=True, color=C_GREEN)
                c.fill = _fill(C_GRN_LT)
            else:
                c.font = _font(bold=has_data)
                c.fill = _fill(C_BLUE_LT if has_data else bg)

    total_r = DATA_START + len(rows)
    _set_row_height(ws, total_r, 28)
    t_rev = sum(r[1] for r in rows)
    t_exp = sum(r[2] for r in rows)
    t_net = t_rev - t_exp

    ws.cell(total_r, 2, "TOTAL").fill = _fill(C_DARK)
    ws.cell(total_r, 2).font      = _font(bold=True, size=11, color=C_WHITE)
    ws.cell(total_r, 2).alignment = _align("center")

    for ci, val, clr in [
        (3, t_rev, C_GOLD),
        (4, t_exp, C_GOLD),
        (5, t_net, C_GREEN if t_net >= 0 else "DC2626"),
    ]:
        c = ws.cell(total_r, ci, val)
        c.number_format = FMT_RP
        c.font      = _font(bold=True, size=11, color=clr)
        c.fill      = _fill(C_DARK)
        c.alignment = _align("right")

    _set_row_height(ws, total_r + 2, 8)
    _section_header(ws, total_r + 3, 2, 5, "📌  Keterangan Warna", height=24)
    legends = [
        (C_BLUE_LT, "Baris dengan data transaksi"),
        (C_GRN_LT,  "Laba bersih positif (untung)"),
        (C_RED_LT,  "Laba bersih negatif (rugi)"),
    ]
    for j, (bg, lbl) in enumerate(legends):
        r = total_r + 4 + j
        _set_row_height(ws, r, 20)
        ws.merge_cells(f"B{r}:E{r}")
        c = ws[f"B{r}"]
        c.value     = "  " + lbl
        c.font      = _font(size=9, color=C_GRAY)
        c.fill      = _fill(bg)
        c.alignment = _align("left")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: 📋 RINGKASAN
# ═══════════════════════════════════════════════════════════════════════════════

def _write_ringkasan(wb, period_label: str, generated_at: str,
                      orders_qs, expenses_qs):
    ws = wb[SH_RING]
    _set_col_widths(ws, {"A": 2, "B": 30, "C": 2, "D": 26, "E": 2})

    _banner(ws, 1, 2, 4, "📋  RINGKASAN KEUANGAN", bg=C_GREEN)

    _set_row_height(ws, 2, 22)
    ws.merge_cells("B2:D2")
    sub = ws["B2"]
    sub.value = f"Periode: {period_label}  |  Digenerate: {generated_at}"
    sub.font  = _font(italic=True, size=10, color=C_GRAY)
    sub.fill  = _fill(C_LIGHT)
    sub.alignment = _align("left")

    _set_row_height(ws, 3, 10)

    paid_orders   = orders_qs.filter(payment_status="paid")
    total_rev     = float(paid_orders.aggregate(t=Sum("total_price"))["t"] or 0)
    count_paid    = paid_orders.count()
    count_pending = orders_qs.exclude(payment_status="paid").count()
    avg_order     = (total_rev / count_paid) if count_paid else 0
    total_exp     = float(expenses_qs.aggregate(t=Sum("amount"))["t"] or 0)
    net           = total_rev - total_exp
    margin        = (net / total_rev * 100) if total_rev else 0

    count_cash = orders_qs.filter(payment_method="cash").count()
    count_qris = orders_qs.filter(payment_method__icontains="qris").count()
    count_web  = orders_qs.filter(source="web").count()
    count_pos  = orders_qs.filter(source="pos").count()

    kpis = [
        ("💰 PENDAPATAN (LUNAS)", f"Rp {int(total_rev):,}".replace(",", "."),
         "Total pendapatan dari order lunas", C_GREEN, C_GRN_LT),
        ("🛒 PENGELUARAN", f"Rp {int(total_exp):,}".replace(",", "."),
         "Total biaya operasional periode ini", C_RED, C_RED_LT),
        ("📈 LABA BERSIH", f"Rp {int(net):,}".replace(",", "."),
         "Pendapatan dikurangi pengeluaran",
         C_GREEN if net >= 0 else "DC2626",
         C_GRN_LT if net >= 0 else C_RED_LT),
    ]

    for i, (title, val_text, sub_text, accent, bg_lt) in enumerate(kpis):
        base = 4 + i * 6
        _set_row_height(ws, base,     5)
        _set_row_height(ws, base + 1, 22)
        _set_row_height(ws, base + 2, 44)
        _set_row_height(ws, base + 3, 20)
        _set_row_height(ws, base + 4, 8)

        ws.merge_cells(f"B{base}:D{base}")
        ws[f"B{base}"].fill = _fill(accent)

        ws.merge_cells(f"B{base+1}:D{base+1}")
        tc = ws[f"B{base+1}"]
        tc.value = title; tc.font = _font(bold=True, size=11, color=accent)
        tc.fill  = _fill(bg_lt); tc.alignment = _align("left")

        ws.merge_cells(f"B{base+2}:D{base+2}")
        vc = ws[f"B{base+2}"]
        vc.value = val_text
        vc.font  = Font(bold=True, size=22, color=accent, name="Arial Black")
        vc.fill  = _fill(bg_lt); vc.alignment = _align("center")

        ws.merge_cells(f"B{base+3}:D{base+3}")
        sc = ws[f"B{base+3}"]
        sc.value = sub_text; sc.font = _font(italic=True, size=9, color=C_GRAY)
        sc.fill  = _fill(bg_lt); sc.alignment = _align("center")

    MET = 4 + 3 * 6 + 1
    _section_header(ws, MET, 2, 4, "📊  Detail Metrik")
    metrics = [
        ("Jumlah Order Lunas",        f"{count_paid} order"),
        ("Jumlah Order Pending",       f"{count_pending} order"),
        ("Rata-rata Nilai per Order",  f"Rp {int(avg_order):,}".replace(",", ".")),
        ("Margin Laba",               f"{margin:.2f}%"),
        ("Transaksi Cash",            f"{count_cash} transaksi"),
        ("Transaksi QRIS",            f"{count_qris} transaksi"),
        ("Order dari Web",            f"{count_web} order"),
        ("Order dari POS",            f"{count_pos} order"),
    ]
    for i, (lbl, val) in enumerate(metrics):
        r  = MET + 1 + i
        bg = C_WHITE if i % 2 == 0 else C_STRIPE
        _set_row_height(ws, r, 24)
        is_neg    = val.startswith("-") or val.startswith("Rp -")
        val_color = "DC2626" if is_neg else (C_GREEN if "Rp" in val and not is_neg else C_BLUE)

        lc = ws.cell(r, 2, lbl)
        lc.font = _font(size=10); lc.fill = _fill(bg)
        lc.alignment = _align("left"); lc.border = _border_thin()

        vc = ws.cell(r, 4, val)
        vc.font = _font(bold=True, size=10, color=val_color)
        vc.fill = _fill(bg); vc.alignment = _align("right"); vc.border = _border_thin()

    GUIDE = MET + 1 + len(metrics) + 2
    _section_header(ws, GUIDE, 2, 4, "💡  Cara Membaca Laporan Ini")
    guides = [
        "📈 Analisis Tren  →  Perbandingan performa vs periode sebelumnya",
        "💳 Metode Pembayaran  →  Cash vs QRIS vs metode lain, mana dominan",
        "📊 Rekap Periode  →  Pendapatan & pengeluaran per hari atau per bulan",
        "📄 Detail Transaksi  →  Semua order dengan status, metode, dan nominal",
        "🍜 Top Menu  →  Menu terlaris berdasarkan qty dan omzet",
        "💰 Pengeluaran  →  Detail biaya operasional per kategori",
    ]
    for i, text in enumerate(guides):
        r  = GUIDE + 1 + i
        bg = C_LIGHT if i % 2 == 0 else C_WHITE
        _set_row_height(ws, r, 22)
        ws.merge_cells(f"B{r}:D{r}")
        c = ws[f"B{r}"]
        c.value = "  " + text; c.font = _font(size=10)
        c.fill  = _fill(bg);   c.alignment = _align("left")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6: 📄 DETAIL TRANSAKSI
# ═══════════════════════════════════════════════════════════════════════════════

def _write_detail_transaksi(wb, orders_qs, period_label: str, generated_at: str):
    ws = wb[SH_DETAIL]
    _set_col_widths(ws, {
        "A": 5, "B": 20, "C": 14, "D": 8, "E": 14,
        "F": 14, "G": 14, "H": 11, "I": 14, "J": 11,
        "K": 15, "L": 9, "M": 18,
    })

    orders_list = list(orders_qs.order_by("-created_at"))
    _banner(ws, 1, 1, 13, "📄  DETAIL TRANSAKSI", bg=C_ORANGE)

    _set_row_height(ws, 2, 22)
    ws.merge_cells("A2:M2")
    sub = ws["A2"]
    sub.value = (
        f"Periode: {period_label}  |  {len(orders_list)} transaksi"
        f"  |  Digenerate: {generated_at}"
    )
    sub.font  = _font(italic=True, size=10, color=C_GRAY)
    sub.fill  = _fill(C_LIGHT); sub.alignment = _align("left")

    _set_row_height(ws, 3, 8)

    headers = ["No","No. Order","Tanggal","Waktu","Nama","HP",
               "Subtotal","Diskon","Total","Status","Metode","Source","Catatan"]
    _table_header(ws, 4, headers, start_col=1)

    STATUS_STYLE = {
        "paid":    (C_GRN_LT, C_GREEN),
        "Lunas":   (C_GRN_LT, C_GREEN),
        "pending": (C_YLW_LT, C_ORANGE),
        "unpaid":  (C_YLW_LT, C_ORANGE),
        "Pending": (C_YLW_LT, C_ORANGE),
    }

    DATA_START = 5
    for idx, order in enumerate(orders_list, 1):
        r  = DATA_START + idx - 1
        _set_row_height(ws, r, 22)
        status_raw = order.payment_status
        status_lbl = _map_status(status_raw)
        row_bg, status_fg = STATUS_STYLE.get(status_raw, (C_STRIPE, C_GRAY))

        data = [
            (1,  idx,                                        False, FMT_INT, "center"),
            (2,  order.order_number,                         True,  None,    "left"),
            (3,  order.created_at.strftime("%d/%m/%Y"),      False, None,    "center"),
            (4,  order.created_at.strftime("%H:%M"),         False, None,    "center"),
            (5,  order.customer_name or "—",                 False, None,    "left"),
            (6,  order.customer_phone or "—",                False, None,    "left"),
            (7,  float(order.subtotal),                      False, FMT_RP,  "right"),
            (8,  float(order.discount_amount),               False, FMT_RP,  "right"),
            (9,  float(order.total_price),                   True,  FMT_RP,  "right"),
            (10, status_lbl,                                 True,  None,    "center"),
            (11, order.payment_method or "—",                False, None,    "center"),
            (12, _map_source(order.source),                  False, None,    "center"),
            (13, order.notes or "",                          False, None,    "left"),
        ]
        for col_n, val, bold, fmt, align_h in data:
            c = ws.cell(r, col_n, val)
            c.fill      = _fill(row_bg)
            c.alignment = _align(align_h)
            c.border    = _border_bottom()
            c.font      = _font(bold=bold, color=status_fg if col_n == 10 else C_DARK)
            if fmt:
                c.number_format = fmt

    last_r = DATA_START + len(orders_list)
    _set_row_height(ws, last_r + 1, 26)
    paid_total = sum(float(o.total_price) for o in orders_list if o.payment_status == "paid")
    count_paid = sum(1 for o in orders_list if o.payment_status == "paid")
    count_pend = sum(1 for o in orders_list if o.payment_status != "paid")

    ws.merge_cells(f"A{last_r+1}:F{last_r+1}")
    sc = ws[f"A{last_r+1}"]
    sc.value = (
        f"  Total: {len(orders_list)} transaksi"
        f"  |  Lunas: {count_paid}"
        f"  |  Pending: {count_pend}"
    )
    sc.font  = _font(bold=True, size=10, color=C_WHITE)
    sc.fill  = _fill(C_DARK2); sc.alignment = _align("left")

    for ci in range(7, 14):
        ws.cell(last_r + 1, ci).fill = _fill(C_DARK2)

    ws.cell(last_r + 1, 8, "Total Omzet Lunas:").font = _font(bold=True, color=C_GOLD)
    ws.cell(last_r + 1, 8).fill = _fill(C_DARK2)
    ws.cell(last_r + 1, 8).alignment = _align("right")

    tc = ws.cell(last_r + 1, 9, paid_total)
    tc.number_format = FMT_RP
    tc.font = _font(bold=True, size=11, color=C_GOLD)
    tc.fill = _fill(C_DARK2); tc.alignment = _align("right")

    if len(orders_list):
        ws.auto_filter.ref = f"A4:M{DATA_START + len(orders_list) - 1}"


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7: 🍜 TOP MENU
# ═══════════════════════════════════════════════════════════════════════════════

def _write_top_menu(wb, orders_qs, days_in_period: int,
                    period_label: str, generated_at: str):
    ws = wb[SH_TOPMENU]
    _set_col_widths(ws, {"A": 2, "B": 7, "C": 28, "D": 14, "E": 12, "F": 14, "G": 16, "H": 2})

    _banner(ws, 1, 2, 7, "🍜  TOP MENU — TERLARIS", bg="DC2626")

    _set_row_height(ws, 2, 22)
    ws.merge_cells("B2:G2")
    sub = ws["B2"]
    sub.value = f"Berdasarkan qty terjual  |  Periode: {period_label}  |  Digenerate: {generated_at}"
    sub.font  = _font(italic=True, size=10, color=C_GRAY)
    sub.fill  = _fill(C_LIGHT); sub.alignment = _align("left")

    _set_row_height(ws, 3, 8)

    menus = (
        OrderItem.objects.filter(order__in=orders_qs)
        .values("menu__name")
        .annotate(
            qty=Sum("quantity"),
            omzet=Sum(
                ExpressionWrapper(F("price") * F("quantity"), output_field=DecimalField())
            ),
        )
        .order_by("-qty")
    )
    menus_list  = list(menus)
    total_omzet = sum(float(m["omzet"] or 0) for m in menus_list)
    total_qty   = sum(m["qty"] or 0 for m in menus_list)

    headers = ["Rank","Nama Menu","Omzet (Rp)","Qty","% Kontribusi","Rata-rata/hari"]
    _table_header(ws, 4, headers, start_col=2, height=26)
    ws.freeze_panes = None

    RANK_ICONS = ["🥇","🥈","🥉"]
    RANK_BGS   = [C_YLW_LT, C_STRIPE, C_RED_LT]

    DATA_START = 5
    for i, item in enumerate(menus_list):
        r   = DATA_START + i
        _set_row_height(ws, r, 26)
        rank  = RANK_ICONS[i] if i < 3 else str(i + 1)
        omzet = float(item["omzet"] or 0)
        qty   = item["qty"] or 0
        pct   = omzet / total_omzet if total_omzet else 0
        avg   = round(qty / days_in_period, 1) if days_in_period else 0
        bg    = RANK_BGS[i] if i < 3 else (C_STRIPE if i % 2 == 0 else C_WHITE)

        data = [
            (2, rank,               True  if i == 0 else False, None,    "center"),
            (3, item["menu__name"], True  if i == 0 else False, None,    "left"),
            (4, omzet,              False, FMT_RP,  "right"),
            (5, qty,                False, FMT_INT, "center"),
            (6, pct,                False, FMT_PCT, "right"),
            (7, avg,                False, "0.0",   "center"),
        ]
        for col_n, val, bold, fmt, align_h in data:
            c = ws.cell(r, col_n, val)
            c.fill      = _fill(bg)
            c.alignment = _align(align_h)
            c.border    = _border_bottom()
            fg = C_GOLD if i == 0 else C_DARK
            c.font      = _font(bold=bold, color=fg)
            if fmt:
                c.number_format = fmt

    total_r = DATA_START + len(menus_list)
    _set_row_height(ws, total_r, 26)
    for col in range(2, 8):
        ws.cell(total_r, col).fill = _fill(C_DARK)
    ws.cell(total_r, 2, "TOTAL").font = _font(bold=True, size=11, color=C_WHITE)
    ws.cell(total_r, 2).alignment = _align("center")

    ws.cell(total_r, 4, total_omzet).number_format = FMT_RP
    ws.cell(total_r, 4).font = _font(bold=True, size=11, color=C_GOLD)
    ws.cell(total_r, 4).fill = _fill(C_DARK); ws.cell(total_r, 4).alignment = _align("right")

    ws.cell(total_r, 5, total_qty).font = _font(bold=True, size=11, color=C_GOLD)
    ws.cell(total_r, 5).fill = _fill(C_DARK); ws.cell(total_r, 5).alignment = _align("center")

    ws.cell(total_r, 6, 1.0).number_format = FMT_PCT
    ws.cell(total_r, 6).font = _font(bold=True, size=11, color=C_GOLD)
    ws.cell(total_r, 6).fill = _fill(C_DARK); ws.cell(total_r, 6).alignment = _align("right")

    if menus_list:
        best_qty = max(menus_list, key=lambda x: x["qty"] or 0)
        best_rev = max(menus_list, key=lambda x: float(x["omzet"] or 0))
        ins_r = total_r + 2
        _section_header(ws, ins_r, 2, 7, "💡  Insight Otomatis")
        insights = [
            f"Menu terlaris (qty): {best_qty['menu__name']} — {best_qty['qty']} pcs terjual",
            f"Menu omzet tertinggi: {best_rev['menu__name']} — Rp {int(float(best_rev['omzet'] or 0)):,}".replace(",", "."),
            f"Total menu terjual: {total_qty} pcs  |  Total omzet: Rp {int(total_omzet):,}".replace(",", "."),
        ]
        for j, text in enumerate(insights):
            r  = ins_r + 1 + j
            _set_row_height(ws, r, 22)
            ws.merge_cells(f"B{r}:G{r}")
            c = ws[f"B{r}"]
            c.value = "  ▸  " + text
            c.font  = _font(size=10)
            c.fill  = _fill(C_LIGHT if j % 2 == 0 else C_WHITE)
            c.alignment = _align("left")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8: 💰 PENGELUARAN
# ═══════════════════════════════════════════════════════════════════════════════

def _write_pengeluaran(wb, expenses_qs, period_label: str, generated_at: str):
    ws = wb[SH_EXPENSE]
    _set_col_widths(ws, {"A": 2, "B": 6, "C": 14, "D": 18, "E": 28, "F": 18, "G": 14, "H": 22, "I": 2})

    expenses_list = list(expenses_qs.order_by("date"))
    _banner(ws, 1, 2, 8, "💰  PENGELUARAN", bg=C_PURPLE)

    _set_row_height(ws, 2, 22)
    ws.merge_cells("B2:H2")
    sub = ws["B2"]
    sub.value = (
        f"Periode: {period_label}  |  {len(expenses_list)} item"
        f"  |  Digenerate: {generated_at}"
    )
    sub.font  = _font(italic=True, size=10, color=C_GRAY)
    sub.fill  = _fill(C_LIGHT); sub.alignment = _align("left")

    _set_row_height(ws, 3, 8)

    headers = ["No","Tanggal","Kategori","Keterangan","Nominal (Rp)","Metode","Catatan"]
    _table_header(ws, 4, headers, start_col=2, height=26)

    CAT_STYLE = {
        "Operasional": (C_YLW_LT,  C_ORANGE),
        "Bahan Baku":  (C_GRN_LT,  C_GREEN),
        "Karyawan":    (C_BLUE_LT, C_BLUE),
        "Marketing":   ("F3E8FF",  C_PURPLE),
    }

    DATA_START = 5
    for i, exp in enumerate(expenses_list):
        r   = DATA_START + i
        _set_row_height(ws, r, 22)
        cat  = getattr(exp, "category", "Operasional")
        bg, fg = CAT_STYLE.get(cat, (C_STRIPE, C_DARK))

        data = [
            (2, i + 1,                                  False, None,   "center"),
            (3, exp.date.strftime("%d/%m/%Y"),           False, None,   "center"),
            (4, cat,                                     True,  None,   "left"),
            (5, exp.description,                         False, None,   "left"),
            (6, float(exp.amount),                       True,  FMT_RP, "right"),
            (7, getattr(exp, "payment_method", "Cash"),  False, None,   "center"),
            (8, getattr(exp, "notes", "") or "",         False, None,   "left"),
        ]
        for col_n, val, bold, fmt, align_h in data:
            c = ws.cell(r, col_n, val)
            c.fill      = _fill(bg)
            c.alignment = _align(align_h)
            c.border    = _border_bottom()
            c.font      = _font(bold=bold, color=fg if col_n == 4 else C_DARK)
            if fmt:
                c.number_format = fmt

    total_exp   = float(expenses_qs.aggregate(t=Sum("amount"))["t"] or 0)
    last_data_r = DATA_START + len(expenses_list) - 1
    total_r     = last_data_r + 2
    _set_row_height(ws, total_r, 28)

    ws.merge_cells(f"B{total_r}:E{total_r}")
    tc = ws[f"B{total_r}"]
    tc.value = "  TOTAL PENGELUARAN"
    tc.font  = _font(bold=True, size=11, color=C_WHITE)
    tc.fill  = _fill(C_DARK); tc.alignment = _align("left")

    vc = ws.cell(total_r, 6, total_exp)
    vc.number_format = FMT_RP
    vc.font      = _font(bold=True, size=12, color=C_GOLD)
    vc.fill      = _fill(C_DARK); vc.alignment = _align("right")
    for col in [7, 8]:
        ws.cell(total_r, col).fill = _fill(C_DARK)

    rekap_r = total_r + 3
    _section_header(ws, rekap_r, 2, 8, "📊  REKAP PENGELUARAN PER KATEGORI")

    cat_headers = ["Kategori", "Total (Rp)", "% dari Total", "Status"]
    _set_row_height(ws, rekap_r + 1, 24)
    for i, (col_n, h) in enumerate(zip([2, 6, 8, 4], cat_headers)):
        c = ws.cell(rekap_r + 1, col_n, h)
        c.font = _font(bold=True, color=C_WHITE)
        c.fill = _fill(C_DARK2); c.alignment = _align("center")

    try:
        db_cats = list(expenses_qs.values_list("category", flat=True).distinct())
    except Exception:
        db_cats = []
    all_cats = list(dict.fromkeys(db_cats + ["Bahan Baku", "Operasional", "Karyawan", "Marketing"]))

    cat_totals = {}
    for exp in expenses_list:
        cat = getattr(exp, "category", "Operasional")
        cat_totals[cat] = cat_totals.get(cat, 0) + float(exp.amount)

    for i, cat in enumerate(all_cats):
        r = rekap_r + 2 + i
        _set_row_height(ws, r, 22)
        bg, fg = CAT_STYLE.get(cat, (C_STRIPE, C_DARK))
        amt = cat_totals.get(cat, 0)
        pct = amt / total_exp if total_exp and amt else 0

        ws.merge_cells(f"B{r}:E{r}")
        lc = ws[f"B{r}"]
        lc.value = cat; lc.font = _font(bold=True, color=fg)
        lc.fill = _fill(bg); lc.alignment = _align("left")

        vc = ws.cell(r, 6, amt)
        vc.number_format = FMT_RP
        vc.font = _font(bold=True, color=C_GREEN if amt == 0 else C_DARK)
        vc.fill = _fill(bg); vc.alignment = _align("right")

        pc = ws.cell(r, 8, pct)
        pc.number_format = FMT_PCT
        pc.font = _font(color=C_GRAY)
        pc.fill = _fill(bg); pc.alignment = _align("right")


# ═══════════════════════════════════════════════════════════════════════════════
# FUNGSI UTAMA
# ═══════════════════════════════════════════════════════════════════════════════

def export_finance_excel(
    request,
    orders_qs,
    expenses_qs,
    period_label: str,
    mode: str = "monthly",
    month: int | None = None,
    year: int | None = None,
) -> HttpResponse:
    """
    Ekspor laporan keuangan ke Excel.

    Sheet order: Cover → Analisis Tren → Metode Pembayaran → Rekap Periode
                 → Ringkasan → Detail Transaksi → Top Menu → Pengeluaran

    Parameters
    ----------
    request       : Django HttpRequest
    orders_qs     : QuerySet Order (sudah difilter periode)
    expenses_qs   : QuerySet Expense (sudah difilter periode)
    period_label  : Label teks, mis. 'Juni 2026' / 'Tahun 2026'
    mode          : 'monthly' | 'yearly'
    month         : 1–12, wajib jika mode='monthly'
    year          : 4-digit year
    """
    now = timezone.now()
    if year is None:
        year = now.year
    if mode == "monthly" and month is None:
        month = now.month

    generated_at   = _now_label()
    filename       = _build_filename(mode, month, year)
    days_in_period = calendar.monthrange(year, month)[1] if mode == "monthly" else 365

    total_rev = float(
        orders_qs.filter(payment_status="paid")
        .aggregate(t=Sum("total_price"))["t"] or 0
    )
    total_exp = float(expenses_qs.aggregate(t=Sum("amount"))["t"] or 0)
    total_net = total_rev - total_exp

    wb = _create_workbook()

    _write_analisis_tren(wb, mode, month, year, orders_qs, expenses_qs, generated_at)
    _write_rekap_metode_pembayaran(wb, orders_qs, period_label, generated_at)
    _write_rekap_periode(wb, mode, month, year, orders_qs, expenses_qs, generated_at)
    _write_ringkasan(wb, period_label, generated_at, orders_qs, expenses_qs)
    _write_detail_transaksi(wb, orders_qs, period_label, generated_at)
    _write_top_menu(wb, orders_qs, days_in_period, period_label, generated_at)
    _write_pengeluaran(wb, expenses_qs, period_label, generated_at)
    _write_cover(wb, period_label, generated_at, mode, year,
                 total_rev, total_exp, total_net)

    wb.active = wb[SH_RING]

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    encoded  = filename.encode("utf-8").decode("latin-1", errors="replace")
    response["Content-Disposition"] = (
        f'attachment; filename="{encoded}"; '
        f"filename*=UTF-8''{filename.replace(' ', '%20')}"
    )
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════════════════════
# VIEW HELPER
# ═══════════════════════════════════════════════════════════════════════════════

def export_finance_excel_view(request) -> HttpResponse:
    """
    GET /api/orders/export/finance-excel/?mode=monthly&month=6&year=2026
    GET /api/orders/export/finance-excel/?mode=yearly&year=2026
    """
    now   = timezone.now()
    mode  = request.GET.get("mode", "monthly")
    year  = int(request.GET.get("year",  now.year))
    month = int(request.GET.get("month", now.month)) if mode == "monthly" else None

    if mode == "monthly":
        orders_qs   = Order.objects.filter(
            created_at__year=year, created_at__month=month
        ).prefetch_related("items__menu")
        expenses_qs = Expense.objects.filter(date__year=year, date__month=month)
        period_label = f"{BULAN_ID[month]} {year}"
    else:
        orders_qs   = Order.objects.filter(created_at__year=year).prefetch_related("items__menu")
        expenses_qs = Expense.objects.filter(date__year=year)
        period_label = f"Tahun {year}"

    return export_finance_excel(
        request=request,
        orders_qs=orders_qs,
        expenses_qs=expenses_qs,
        period_label=period_label,
        mode=mode,
        month=month,
        year=year,
    )